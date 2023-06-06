from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import load_workbook


driver = webdriver.Chrome()
driver.get("https://testing.wenodo.com/")
driver.maximize_window()

# Login
email_input = driver.find_element(By.XPATH, "//input[@placeholder='E-mail Address']")
email_input.send_keys("mubashir@gmail.com")

password_input = driver.find_element(By.XPATH, "//input[@id='password']")
password_input.send_keys("demo123")

login_button = driver.find_element(By.XPATH, "//button[@id='submit']")
login_button.click()



#click on human resource
time.sleep(5)
driver.implicitly_wait(20)  # Click on Human Resource
hr_button = driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div/div[2]/div[2]/div[2]/a/span")
hr_button.click()

# Load Excel file
workbook = load_workbook("C:\\Users\\sharad\\Downloads\\DEPARTMENT_DEMO WENODO UK_16_2_6_2023_300.xlsx")
sheet = workbook.active

# Iterate through rows starting from 2nd row
for row in sheet.iter_rows(min_row=2, values_only=True):
    department_name = row[0]
    division_name = row[1]
    reporting_department = row[2]

    # Click on Add button
    add_button = driver.find_element(By.XPATH,
                                     "/html[1]/body[1]/div[3]/div[1]/section[1]/div[1]/ng-view[1]/div[2]/div[1]/div[2]/div[1]/div[1]/button[1]")
    add_button.click()
    time.sleep(5)
    driver.implicitly_wait(20)

    # Click on Add department
    add_department_button = driver.find_element(By.XPATH,
                                                "/html/body/div[3]/div/section/div/ng-view/div[2]/div[1]/div[2]/div/div/div[2]/a[1]")
    add_department_button.click()


    time.sleep(3)

    # Fill in department name
    department_name_input = driver.find_element(By.XPATH,
                                                "/html/body/div[3]/div/section/div/ng-view/form/div/div[2]/templateurl/div/div[2]/div[1]/div/div/input")
    department_name_input.send_keys(department_name)

    time.sleep(2)
    driver.implicitly_wait(20)

    # Select division name
    division_name_select = driver.find_element(By.XPATH, "//select[@name='FIELD_NAME302']")
    division_name_select.send_keys(division_name)

    time.sleep(2)

    # Select reporting department
    reporting_department_select = driver.find_element(By.XPATH, "//select[@name='FIELD_NAME503']")
    reporting_department_select.click()
    reporting_department_select.send_keys(reporting_department)
    reporting_department_select.click()

    time.sleep(1)







    # Submit the form
    submit_button = driver.find_element(By.XPATH, "/html/body/div[3]/div/section/div/ng-view/form/div/div[4]/div/button[1]")
    submit_button.click()
    time.sleep(5)

    # Go back to the home page
    home_button = driver.find_element(By.XPATH, "//img[@class='brand-image']")
    home_button.click()

    time.sleep(5)

driver.quit()
